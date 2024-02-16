import openpyxl


def getRow_count(path, sheetName):
    work_book = openpyxl.load_workbook(path)
    sheet = work_book[sheetName]
    return sheet.max_row


def getColumn_count(path, sheetName):
    work_book = openpyxl.load_workbook(path)
    sheet = work_book[sheetName]
    return sheet.max_column


def read_data(path, sheetName, r, c):
    work_book = openpyxl.load_workbook(path)
    sheet = work_book[sheetName]
    return sheet.cell(row=r, column=c).value


def write_data(path, sheetName, r, c, data):
    work_book = openpyxl.load_workbook(path)
    sheet = work_book[sheetName]
    sheet.cell(row=r, column=c).value = data
    work_book.save(path)
