import openpyxl

path = "C:\\Users\\MyPc\\Desktop\\DDF.xlsx"

wb=openpyxl.load_workbook(path)
#sh=wb.get_sheet_by_name('Sheet1')
sh=wb['Sheet1']
rows=sh.max_row
print('total rows in excelsheet=' , rows)
col=sh.max_column
print("total columns in excelsheet =", col)

for r in range(1,rows+1):
    for c in range(1, col+1):
        print(sh.cell(row=r, column=c).value, end=" ")

print()